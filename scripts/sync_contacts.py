"""
EnviroCare — extract Contact Log entries (READ-ONLY by default)

Usage:
    FIELDSTER_API_TOKEN=xxx python scripts/sync_contacts.py <tracker.xlsx> [--csv OUT.csv] [--push]

WHAT IT DOES
    Reads the "Contact Log" tab of the tracker workbook and reports every
    collections call/text that would be synced to a customer's Fieldster
    record. By default it only EXTRACTS and prints (and optionally writes a
    CSV) — it makes no external changes.

SAFETY / GUARDRAILS  (read before using --push)
    This tool NEVER modifies customer contacts or billing by default.
    It does not POST to Fieldster and does not write back to the workbook
    unless writes are explicitly authorized for that run.

    Pushing entries to Fieldster writes to live customer records. It is
    DISABLED unless BOTH of these "specific safeguards" are present:
        1. the  --push  command-line flag, AND
        2. the  FIELDSTER_ALLOW_WRITES=1  environment variable.
    Missing either one => the script refuses to push and stays read-only.
    Do not enable writes to customer contacts or billing without explicit,
    per-run authorization. The API details below are also still unverified
    (marked TODO), so confirm them before ever turning writes on.

    The credential is read from FIELDSTER_API_TOKEN (matching .env.example),
    with FIELDSTER_API_KEY accepted as a fallback. Never hard-code the key in
    this file — this is a public repository.
"""

import sys, os, csv, datetime

from openpyxl import load_workbook


BASE_URL = "https://api.fieldster.com"          # TODO: confirm from API docs


def build_payload(fid, date, channel, who, outcome, notes):
    """Shape a Contact Log row into the Fieldster recordcall payload."""
    return {
        "customer_id": str(fid).replace("F-", ""),  # TODO: confirm ID format
        "call_direction": "outbound",
        "disposition": f"Collections - {outcome}",
        "notes": f"[{channel}] {notes}".strip(),
        "agent_name": who or "EnviroCare Office",
        "date_started": str(date),
        "date_completed": str(date),
    }


def extract(xlsx_path):
    """Read the Contact Log tab and return the entries eligible to sync.

    Read-only: opens the workbook but never saves it.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    lg = wb["Contact Log"]
    entries = []
    for row in range(2, 1002):
        fid = lg.cell(row=row, column=2).value
        if not fid: continue
        notes = str(lg.cell(row=row, column=7).value or "")
        if "[SYNCED]" in notes or "SAMPLE" in notes.upper(): continue
        date = lg.cell(row=row, column=1).value
        channel = lg.cell(row=row, column=3).value or ""
        who = lg.cell(row=row, column=5).value or ""
        outcome = lg.cell(row=row, column=6).value or ""
        if isinstance(date, datetime.datetime): date = date.date()
        entries.append({
            "row": row,
            "fieldster_id": fid,
            "payload": build_payload(fid, date, channel, who, outcome, notes),
        })
    wb.close()
    return entries


def write_csv(entries, csv_path):
    fields = ["row", "customer_id", "date_started", "call_direction",
              "disposition", "agent_name", "notes"]
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for e in entries:
            p = e["payload"]
            w.writerow({"row": e["row"], **{k: p[k] for k in fields if k in p}})
    print(f"Wrote {len(entries)} entr(ies) to {csv_path}")


def push(xlsx_path, entries):
    """Write each entry to Fieldster. Guarded by extract-then-authorize logic
    in main(); only called once both safeguards are confirmed present."""
    import requests  # imported lazily — not needed for read-only extraction
    api_key = os.environ.get("FIELDSTER_API_TOKEN") or os.environ.get("FIELDSTER_API_KEY")
    if not api_key:
        sys.exit("Missing credential: set FIELDSTER_API_TOKEN (or FIELDSTER_API_KEY).")
    headers = {"Authorization": f"Bearer {api_key}"}  # TODO: confirm auth scheme
    # Re-open writable so we can mark synced rows.
    wb = load_workbook(xlsx_path)
    lg = wb["Contact Log"]
    synced = 0
    for e in entries:
        r = requests.post(f"{BASE_URL}/api/customers/recordcall",
                          json=e["payload"], headers=headers, timeout=15)
        if r.ok:
            cell = lg.cell(row=e["row"], column=7)
            cell.value = (str(cell.value or "") + " [SYNCED]").strip()
            synced += 1
        else:
            print(f"Row {e['row']} ({e['fieldster_id']}) failed: {r.status_code} {r.text[:120]}")
    wb.save(xlsx_path)
    print(f"Synced {synced} contact(s) to Fieldster.")


def main(xlsx_path, csv_path=None, do_push=False):
    entries = extract(xlsx_path)
    print(f"Extracted {len(entries)} contact(s) eligible to sync (read-only).")
    for e in entries:
        print(f"  Row {e['row']} ({e['fieldster_id']}) -> {e['payload']}")
    if csv_path:
        write_csv(entries, csv_path)

    if not do_push:
        return
    # --- write path: only reachable with BOTH safeguards present ---
    if os.environ.get("FIELDSTER_ALLOW_WRITES") != "1":
        sys.exit(
            "Refusing to push: writing to customer records is disabled. "
            "Both --push AND FIELDSTER_ALLOW_WRITES=1 are required to authorize "
            "a run (see the SAFETY / GUARDRAILS section in this file's header)."
        )
    push(xlsx_path, entries)


if __name__ == "__main__":
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    positional = [a for a in sys.argv[1:] if not a.startswith("--")]
    do_push = "--push" in flags
    csv_path = None
    if "--csv" in sys.argv[1:]:
        i = sys.argv[1:].index("--csv")
        rest = sys.argv[1:][i + 1:]
        if not rest or rest[0].startswith("--"):
            print("--csv requires a file path"); sys.exit(1)
        csv_path = rest[0]
        positional = [p for p in positional if p != csv_path]
    if len(positional) != 1:
        print(__doc__); sys.exit(1)
    main(positional[0], csv_path=csv_path, do_push=do_push)
